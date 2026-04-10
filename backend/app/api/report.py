import os
import tempfile

import pandas as pd
from fastapi import APIRouter, HTTPException
from fastapi.responses import FileResponse
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from starlette.background import BackgroundTask
from datetime import datetime
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

from app.core.tasks import tasks

router = APIRouter()


def format_results(results):
    formatted = []

    for item in results:
        formatted.append(
            {
                "Target": item.get("target", ""),
                "Protocol": item.get("protocol", ""),
                "Username": item.get("username", ""),
                "Password": item.get("password", ""),
                "Status": item.get("status", ""),
            }
        )

    return formatted


@router.get("/export/{task_id}/pdf")
async def export_pdf(task_id: str):

    task_list = tasks.list_serialized_tasks()

    task = next(
        (t for t in task_list
         if t.get("task_id") == task_id or t.get("task_id", "").startswith(task_id)),
        None
    )

    if not task:
        raise HTTPException(status_code=404, detail="Task not found")

    if task.get("status") not in ["completed", "success"]:
        raise HTTPException(status_code=400, detail="Task not completed")

    results = task.get("result") or task.get("results") or []
    formatted = format_results(results)

    # ======================
    # 📊 Statistics
    # ======================
    total = len(results)
    weak_count = len(formatted)
    hit_rate = f"{(weak_count / total * 100):.2f}%" if total else "0%"

    if weak_count == 0:
        risk_level = "Low"
        risk_desc = "No weak passwords detected. System is in good security condition."
    elif weak_count < 5:
        risk_level = "Medium"
        risk_desc = "Some weak passwords detected. Immediate improvement recommended."
    else:
        risk_level = "High"
        risk_desc = "Severe weak password risks detected. System is vulnerable."

    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:

        doc = SimpleDocTemplate(
            tmp.name,
            pagesize=letter,
            rightMargin=30,
            leftMargin=30,
            topMargin=30,
            bottomMargin=20,
        )

        styles = getSampleStyleSheet()
        elements = []

        # ======================
        # 🟦 1. Title
        # ======================
        elements.append(Paragraph("Weak Password Security Report", styles["Title"]))
        elements.append(Spacer(1, 20))

        # ======================
        # 🟦 2. Basic Info
        # ======================
        elements.append(Paragraph("1. Basic Information", styles["Heading2"]))
        elements.append(Spacer(1, 8))

        elements.append(Paragraph(f"Task ID: {task_id}", styles["Normal"]))
        elements.append(Paragraph(f"Scan Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles["Normal"]))
        elements.append(Paragraph(f"Status: {task.get('status')}", styles["Normal"]))
        elements.append(Spacer(1, 12))

        # ======================
        # 🟦 3. Statistics
        # ======================
        elements.append(Paragraph("2. Statistics", styles["Heading2"]))
        elements.append(Spacer(1, 8))

        elements.append(Paragraph(f"Total Targets: {total}", styles["Normal"]))
        elements.append(Paragraph(f"Weak Passwords Found: {weak_count}", styles["Normal"]))
        elements.append(Paragraph(f"Hit Rate: {hit_rate}", styles["Normal"]))
        elements.append(Paragraph(f"Risk Level: {risk_level}", styles["Normal"]))
        elements.append(Spacer(1, 12))

        # ======================
        # 🟦 4. Results
        # ======================
        elements.append(Paragraph("3. Scan Results", styles["Heading2"]))
        elements.append(Spacer(1, 8))

        if not formatted:
            elements.append(Paragraph("No weak passwords detected ✅", styles["Normal"]))
        else:
            data = [["Target", "Protocol", "Username", "Password", "Status"]]

            for row in formatted:
                data.append([
                    row.get("Target", "-"),
                    row.get("Protocol", "-"),
                    row.get("Username", "-"),
                    row.get("Password", "-"),
                    row.get("Status", "-"),
                ])

            table = Table(data, repeatRows=1)

            table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.darkblue),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ]))

            elements.append(table)

        elements.append(Spacer(1, 12))

        # ======================
        # 🟦 5. Risk Assessment
        # ======================
        elements.append(Paragraph("4. Risk Assessment", styles["Heading2"]))
        elements.append(Spacer(1, 8))

        elements.append(Paragraph(f"Risk Level: {risk_level}", styles["Normal"]))
        elements.append(Paragraph(risk_desc, styles["Normal"]))
        elements.append(Spacer(1, 12))

        # ======================
        # 🟦 6. Recommendations
        # ======================
        elements.append(Paragraph("5. Recommendations", styles["Heading2"]))
        elements.append(Spacer(1, 8))

        tips = [
            "Use strong passwords (at least 8 characters with letters, numbers, symbols)",
            "Change passwords regularly",
            "Disable default or weak accounts",
            "Enable Multi-Factor Authentication (MFA)",
            "Limit login attempts to prevent brute-force attacks",
        ]

        for tip in tips:
            elements.append(Paragraph(f"- {tip}", styles["Normal"]))

        # ======================
        # 💾 Generate PDF
        # ======================
        doc.build(elements)

        return FileResponse(
            tmp.name,
            media_type="application/pdf",
            filename=f"weak_password_report_{task_id}.pdf",
            background=BackgroundTask(os.unlink, tmp.name),
        )


@router.get("/export/{task_id}/excel")
async def export_excel(task_id: str):

    task_list = tasks.list_serialized_tasks()

    task = next(
        (t for t in task_list
         if t.get("task_id") == task_id or t.get("task_id", "").startswith(task_id)),
        None
    )

    if not task:
        raise HTTPException(status_code=404, detail="任务不存在")

    if task.get("status") not in ["completed", "success"]:
        raise HTTPException(status_code=400, detail="任务尚未完成")

    results = task.get("result") or task.get("results") or []
    formatted = format_results(results)

    # ======================
    # 📊 统计信息
    # ======================
    total = len(results)
    weak_count = len(formatted)

    hit_rate = f"{(weak_count / total * 100):.2f}%" if total else "0%"

    if weak_count == 0:
        risk_level = "低风险"
    elif weak_count < 5:
        risk_level = "中风险"
    else:
        risk_level = "高风险"

    df = pd.DataFrame(formatted)

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:

        wb = Workbook()

        # ======================
        # 🟦 Sheet1：汇总信息
        # ======================
        ws_summary = wb.active
        ws_summary.title = "汇总信息"

        ws_summary.append(["弱口令检测报告"])
        ws_summary.append([])

        summary_data = [
            ["任务ID", task_id],
            ["扫描时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["任务状态", task.get("status")],
            [],
            ["总目标数", total],
            ["弱口令数量", weak_count],
            ["命中率", hit_rate],
            ["风险等级", risk_level],
        ]

        for row in summary_data:
            ws_summary.append(row)

        # 标题样式
        ws_summary["A1"].font = Font(bold=True, size=16)

        # ======================
        # 🟦 Sheet2：检测结果
        # ======================
        ws_result = wb.create_sheet("检测结果")

        if df.empty:
            ws_result.append(["未发现弱口令"])
        else:
            # 中文表头映射
            ws_result.append(["目标", "协议", "用户名", "密码", "状态"])

            for row in dataframe_to_rows(df, index=False, header=False):
                ws_result.append(row)

            # 表头样式
            for cell in ws_result[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

            # 自动列宽
            for col in ws_result.columns:
                max_len = 0
                col_letter = col[0].column_letter

                for cell in col:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))

                ws_result.column_dimensions[col_letter].width = max_len + 4

        # ======================
        # 🟦 Sheet3：安全建议
        # ======================
        ws_advice = wb.create_sheet("安全建议")

        ws_advice.append(["安全加固建议"])
        ws_advice["A1"].font = Font(bold=True, size=14)

        tips = [
            "使用强密码（建议8位以上，包含字母+数字+符号）",
            "定期更换密码，避免长期使用同一密码",
            "禁用默认账户或弱默认密码",
            "启用多因素认证（MFA）",
            "限制登录尝试次数，防止暴力破解",
        ]

        for tip in tips:
            ws_advice.append([tip])

        # 居左对齐
        for row in ws_advice.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)

        # ======================
        # 💾 保存文件
        # ======================
        wb.save(tmp.name)

        return FileResponse(
            tmp.name,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=f"弱口令检测报告_{task_id}.xlsx",
            background=BackgroundTask(os.unlink, tmp.name),
        )